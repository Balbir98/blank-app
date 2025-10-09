# app.py
import io
import numpy as np
import pandas as pd
import streamlit as st
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

st.set_page_config(page_title="Quarterly MI Builder", layout="wide")
st.title("Quarterly MI Builder (Protection & GI)")

# --------------------------------------
# Provider theme rules (header + titles)
# --------------------------------------
def provider_theme(name: str):
    key = (name or "").strip().lower()
    theme = {"header_fill": "000000", "title_font_color": "FFFFFF"}
    if key == "aviva":
        theme["header_fill"] = "FBDB04"; theme["title_font_color"] = "1151AD"
    elif key in ["cirencester", "circencester", "cirencester friendly"]:
        theme["header_fill"] = "9A268C"; theme["title_font_color"] = "FFFFFF"
    elif key == "guardian":
        # UPDATED: Guardian branding
        theme["header_fill"] = "FFC000"; theme["title_font_color"] = "000000"
    elif key in ["lv", "lv="]:
        theme["header_fill"] = "00B050"; theme["title_font_color"] = "FFFFFF"
    elif key in ["payment shield", "paymentshield", "payment-shield"]:
        theme["header_fill"] = "000000"; theme["title_font_color"] = "FFFFFF"
    return theme

# --------------------
# Data load + cleanup
# --------------------
REQ_COLS = [
    "Adviser Name","Firm Name","Application Date","Holder",
    "Product Type","Product Sub Type","Provider","API","Month","Quarter"
]
MONTH_ORDER = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def coerce_month_order(s: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(s):
        mapper = {i+1: MONTH_ORDER[i] for i in range(12)}
        s = s.map(mapper)
    else:
        s = s.astype(str).str.strip().str[:3].str.title()
        s = s.where(s.isin(MONTH_ORDER), np.nan)
    return pd.Categorical(s, categories=MONTH_ORDER, ordered=True)

def read_table(upload) -> pd.DataFrame:
    if upload.name.lower().endswith(".csv"):
        df = pd.read_csv(upload)
    else:
        df = pd.read_excel(upload)

    missing = [c for c in REQ_COLS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    for col in ["Provider","Firm Name","Product Type","Product Sub Type"]:
        df[col] = df[col].astype(str).str.strip()

    # Clean API to float
    api_str = df["API"].astype(str).str.replace(r"[^0-9.\-]", "", regex=True).replace("", "0")
    df["API"] = pd.to_numeric(api_str, errors="coerce").fillna(0.0)

    df["Month"] = coerce_month_order(df["Month"])
    return df

# --------------------------
# Firm Level (calculations)
# --------------------------
def firm_level_table(df: pd.DataFrame) -> pd.DataFrame:
    d = df[(df["Firm Name"] != "") & (df["Provider"] != "")]
    api = pd.pivot_table(d, index="Firm Name", columns="Provider", values="API",
                         aggfunc="sum", fill_value=0.0)
    cnt = pd.pivot_table(d.assign(_one=1), index="Firm Name", columns="Provider",
                         values="_one", aggfunc="count", fill_value=0)

    # % of ROW total
    api_row_tot = api.sum(axis=1).replace(0, np.nan)
    cnt_row_tot = cnt.sum(axis=1).replace(0, np.nan)
    api_pct = (api.div(api_row_tot, axis=0) * 100).fillna(0)
    cnt_pct = (cnt.div(cnt_row_tot, axis=0) * 100).fillna(0)

    providers = sorted(set(api.columns).union(cnt.columns))
    blocks = []
    for p in providers:
        a = api_pct[p] if p in api_pct.columns else pd.Series(0, index=api_pct.index)
        c = cnt_pct[p] if p in cnt_pct.columns else pd.Series(0, index=cnt_pct.index)
        blocks.append(pd.DataFrame({(p, "API (%)"): a, (p, "Product Count %"): c}))
    merged = pd.concat(blocks, axis=1)
    merged.index.name = "Firm Name"
    merged[("Total", "API (%)")] = merged.xs("API (%)", axis=1, level=1).sum(axis=1)
    merged[("Total", "Product Count %")] = merged.xs("Product Count %", axis=1, level=1).sum(axis=1)
    merged = merged.round(2)

    # Bottom Grand Total row = provider share of grand totals
    api_total = float(api.values.sum()); cnt_total = float(cnt.values.sum())
    api_provider_share = (api.sum(axis=0) / api_total * 100.0) if api_total > 0 else api.sum(axis=0)*0.0
    cnt_provider_share = (cnt.sum(axis=0) / cnt_total * 100.0) if cnt_total > 0 else cnt.sum(axis=0)*0.0
    data = {}
    for p in providers:
        data[(p, "API (%)")] = round(float(api_provider_share.get(p, 0.0)), 2)
        data[(p, "Product Count %")] = round(float(cnt_provider_share.get(p, 0.0)), 2)
    data[("Total", "API (%)")] = 100.0 if api_total > 0 else 0.0
    data[("Total", "Product Count %")] = 100.0 if cnt_total > 0 else 0.0
    gt = pd.DataFrame([data], index=["Grand Total"]); gt.index.name = "Firm Name"
    final = pd.concat([merged, gt])
    return final.reset_index()

# --------------------------
# Network Spread tables (calculations)
# --------------------------
def network_provider_table(df: pd.DataFrame) -> pd.DataFrame:
    d = df[df["Provider"] != ""].copy()
    api_by_p = d.groupby("Provider", dropna=False)["API"].sum().astype(float)
    api_total = float(api_by_p.sum())
    api_pct = (api_by_p / api_total * 100.0) if api_total > 0 else api_by_p * 0.0

    cnt_by_p = d.groupby("Provider", dropna=False)["Product Type"].count().astype(float)
    cnt_total = float(cnt_by_p.sum())
    cnt_pct = (cnt_by_p / cnt_total * 100.0) if cnt_total > 0 else cnt_by_p * 0.0

    out = pd.DataFrame({
        "Provider": api_pct.index,
        "API (%)": api_pct.values.round(2),
        "Product %": cnt_pct.reindex(api_pct.index).values.round(2)
    }).sort_values("API (%)", ascending=False).reset_index(drop=True)

    # Force Grand Total to exactly 100.00%
    out.loc[len(out)] = ["Grand Total", 100.0 if api_total > 0 else 0.0, 100.0 if cnt_total > 0 else 0.0]
    return out

def network_subtype_by_month_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Rows: Product Sub Type; Cols: months present; values: % of grand total (count).
    Adds 'Grand Total' column and row; returns first column named exactly 'Product Sub Type'.
    """
    d = df[df["Provider"] != ""].copy()
    d["Month"] = pd.Categorical(d["Month"], categories=MONTH_ORDER, ordered=True)

    cnt = pd.pivot_table(
        d.assign(_one=1),
        index="Product Sub Type",
        columns="Month",
        values="_one",
        aggfunc="count",
        fill_value=0
    )

    months_present = [m for m in MONTH_ORDER if m in cnt.columns and cnt[m].sum() > 0]
    cnt = cnt.reindex(columns=months_present)

    total = float(cnt.values.sum())
    pct = (cnt / total * 100.0) if total > 0 else cnt * 0.0
    pct = pct.round(2)

    pct = pct.rename_axis("Product Sub Type")  # prevent 'index' on reset_index

    # Row sums as Grand Total column
    pct["Grand Total"] = pct.sum(axis=1).round(2)

    # Column sums as Grand Total row, with 100 in the corner
    col_totals = pct.drop(columns=["Grand Total"]).sum(axis=0).round(2)
    gt_row = {**{m: col_totals.get(m, 0.0) for m in months_present}, "Grand Total": 100.0}
    pct = pd.concat([pct, pd.DataFrame([gt_row], index=["Grand Total"])])

    out = pct[[*months_present, "Grand Total"]].reset_index()
    out.rename(columns={out.columns[0]: "Product Sub Type"}, inplace=True)
    return out

# --------------------------
# Excel writing / formatting
# --------------------------
THIN = Side(style="thin", color="000000")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_df_with_multilevel_header(
    ws: Worksheet, df: pd.DataFrame, start_row: int, start_col: int,
    header_fill="000000", title_font_color="FFFFFF",
    header_font_name="Arial", header_font_size=10,
    body_font_name="Calibri", body_font_size=11,
    add_borders=False, left_align_first_col=True,
    force_percent_from_col=None,  # format all cols from this index as %
):
    """
    - MultiIndex columns → two-row header (provider/metric).
    - Simple columns → single header row.
    - Any column whose header text contains '%' will be written as a percentage.
    - Additionally, if force_percent_from_col is set, columns from that 1-based index onward
      are written as percentages.
    """
    r0, c0 = start_row, start_col
    hdr_fill = PatternFill("solid", fgColor=header_fill)
    hdr_font = Font(name=header_font_name, size=header_font_size, bold=True, color=title_font_color)
    body_font = Font(name=body_font_name, size=body_font_size)

    multi = isinstance(df.columns, pd.MultiIndex)

    if multi:
        cols = list(df.columns)
        first_is_plain = not isinstance(cols[0], tuple)
        multi_cols = cols[1:] if first_is_plain else cols

        col_ptr = c0
        if first_is_plain:
            top = ws.cell(row=r0, column=col_ptr, value="")
            top.font = hdr_font; top.fill = hdr_fill
            top.alignment = Alignment(horizontal="center", vertical="center")
            sub = ws.cell(row=r0+1, column=col_ptr, value=str(cols[0]))
            sub.font = hdr_font; sub.fill = hdr_fill
            sub.alignment = Alignment(horizontal="center", vertical="center")
            col_ptr += 1

        j = 0
        while j < len(multi_cols):
            prov = multi_cols[j][0]
            span = sum(1 for k in range(j, len(multi_cols)) if multi_cols[k][0] == prov)
            ws.cell(row=r0, column=col_ptr, value=str(prov)).font = hdr_font
            ws.cell(row=r0, column=col_ptr).fill = hdr_fill
            ws.cell(row=r0, column=col_ptr).alignment = Alignment(horizontal="center", vertical="center")
            if span > 1:
                ws.merge_cells(start_row=r0, start_column=col_ptr, end_row=r0, end_column=col_ptr+span-1)
            for s in range(span):
                subname = multi_cols[j+s][1]
                cell = ws.cell(row=r0+1, column=col_ptr+s, value=str(subname))
                cell.font = hdr_font; cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            col_ptr += span
            j += span
        data_start = r0 + 2
        cols_to_write = cols
    else:
        # simple header
        for j, name in enumerate(df.columns):
            cell = ws.cell(row=r0, column=c0+j, value=str(name))
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        data_start = r0 + 1
        cols_to_write = list(df.columns)

    # identify percent columns
    first_percent_idx0 = None
    if force_percent_from_col is not None:
        first_percent_idx0 = max(0, int(force_percent_from_col) - 1)

    def is_percent_column(colname, j_idx):
        header_txt = str(colname[1]) if isinstance(colname, tuple) else str(colname)
        by_name = "%" in header_txt
        by_force = (first_percent_idx0 is not None and j_idx >= first_percent_idx0)
        return by_name or by_force

    # body writing (+ percent handling)
    for i in range(len(df)):
        for j, colname in enumerate(cols_to_write):
            v = df.iat[i, j]
            cell = ws.cell(row=data_start+i, column=c0+j)

            if is_percent_column(colname, j):
                try:
                    vfloat = float(v)
                    cell.value = vfloat / 100.0
                    cell.number_format = "0.00%"
                except Exception:
                    cell.value = v
            else:
                cell.value = v

            cell.font = body_font
            cell.alignment = Alignment(horizontal="left" if (left_align_first_col and j == 0) else "center",
                                       vertical="center")

    # last row highlight (footer) — uses same header fill & bold font
    last_row = data_start + len(df) - 1
    for j in range(len(cols_to_write)):
        cell = ws.cell(row=last_row, column=c0+j)
        cell.font = Font(name=header_font_name, size=header_font_size, bold=True, color=title_font_color)
        cell.fill = hdr_fill

    # borders
    if add_borders:
        end_col = c0 + len(cols_to_write) - 1
        for rr in range(r0, last_row+1):
            for cc in range(c0, end_col+1):
                ws.cell(row=rr, column=cc).border = BORDER_ALL

    # widths
    for j, name in enumerate(cols_to_write):
        label = name if not isinstance(name, tuple) else name[1]
        ws.column_dimensions[get_column_letter(c0+j)].width = max(12, len(str(label)) + 2)

# --------------------------
# Helpers (Network Spreads H8 block)
# --------------------------
def _norm(s: str) -> str:
    return "".join(ch for ch in s.lower() if ch.isalnum())

def pick_network_sheet(wb):
    wanted = {_norm("Network Spread"), _norm("Network Spreads")}
    for name in wb.sheetnames:
        if _norm(name) in wanted:
            return wb[name]
    return wb.worksheets[1] if len(wb.worksheets) > 1 else wb.create_sheet("Network Spread")

def _overlaps(a_minr, a_minc, a_maxr, a_maxc, ref: str) -> bool:
    minc, minr, maxc, maxr = range_boundaries(ref)
    return not (a_maxr < minr or a_minr > maxr or a_maxc < minc or a_minc > maxc)

def remove_tables_in_range(ws: Worksheet, min_row: int, min_col: int, max_row: int, max_col: int):
    tables = []
    if hasattr(ws, "tables") and ws.tables:
        tables = list(ws.tables.values()) if isinstance(ws.tables, dict) else list(ws.tables)
    elif hasattr(ws, "_tables"):
        tables = list(ws._tables)
    for t in list(tables):
        try:
            if _overlaps(min_row, min_col, max_row, max_col, t.ref):
                if hasattr(ws, "tables") and isinstance(ws.tables, dict):
                    ws.tables.pop(t.name, None)
                elif hasattr(ws, "_tables"):
                    ws._tables.remove(t)
        except Exception:
            pass

def find_and_fix_header_block(ws: Worksheet, df: pd.DataFrame, theme, default_row=8, default_col=8):
    # locate the header cell near H8
    targets = {"index", "product sub type", "product subtype", "product sub-type", "row labels"}
    row, col, found = default_row, default_col, False
    for r in range(default_row-2, default_row+3):
        for c in range(default_col-2, default_col+3):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower() in targets:
                row, col, found = r, c, True
                break
        if found: break
    if not found:
        for r in range(6, 30):
            for c in range(7, 40):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip().lower() in targets:
                    row, col, found = r, c, True
                    break
            if found: break

    # clear any Excel Table that might lock styles
    rows, cols = df.shape
    remove_tables_in_range(ws, row, col, row + rows, col + cols - 1)

    # enforce header cell text/style
    hdr_font = Font(name="Arial", size=10, bold=True, color=theme["title_font_color"])
    hdr_fill = PatternFill("solid", fgColor=theme["header_fill"])
    hc = ws.cell(row=row, column=col)
    hc.value = "Product Sub Type"
    hc.font = hdr_font
    hc.fill = hdr_fill
    hc.alignment = Alignment(horizontal="center", vertical="center")

    # deterministically write the numeric body from df (value/100, format %)
    first_data_row = row + 1
    for i in range(rows):
        for j in range(1, cols):  # skip first label column
            cell = ws.cell(row=first_data_row + i, column=col + j)
            try:
                v = float(df.iat[i, j])        # e.g., 2.16
                cell.value = v / 100.0         # -> 0.0216
                cell.number_format = "0.00%"
            except Exception:
                pass

# --------------------------
# Build workbook
# --------------------------
def build_workbook(template_bytes: bytes, df: pd.DataFrame, provider_choice: str) -> bytes:
    wb = load_workbook(io.BytesIO(template_bytes))
    ws_firm = wb["Firm Level"] if "Firm Level" in wb.sheetnames else wb.worksheets[0]
    ws_network = pick_network_sheet(wb)
    theme = provider_theme(provider_choice)

    # tables
    t1 = firm_level_table(df)               # Firm Level (percent values)
    t2 = network_provider_table(df)         # Provider share (percent values)
    t3 = network_subtype_by_month_table(df) # Product Sub Type by Month (percent values)

    # Firm Level (A6)
    write_df_with_multilevel_header(
        ws_firm, t1, start_row=6, start_col=1,
        header_fill=theme["header_fill"], title_font_color=theme["title_font_color"],
        add_borders=False, left_align_first_col=True,
        force_percent_from_col=2
    )

    # Network Spreads — Provider table (A8)
    write_df_with_multilevel_header(
        ws_network, t2, start_row=8, start_col=1,
        header_fill=theme["header_fill"], title_font_color=theme["title_font_color"],
        add_borders=True, left_align_first_col=True,
        force_percent_from_col=2
    )

    # Network Spreads — Product Sub Type by Month (H8)
    rows, cols = t3.shape
    out_min_row, out_min_col = 8, 8
    out_max_row = out_min_row + rows
    out_max_col = out_min_col + cols - 1
    remove_tables_in_range(ws_network, out_min_row, out_min_col, out_max_row, out_max_col)

    t3.rename(columns={t3.columns[0]: "Product Sub Type"}, inplace=True)
    write_df_with_multilevel_header(
        ws_network, t3, start_row=8, start_col=8,
        header_fill=theme["header_fill"], title_font_color=theme["title_font_color"],
        add_borders=True, left_align_first_col=True,
        force_percent_from_col=2
    )
    find_and_fix_header_block(ws_network, t3, theme, default_row=8, default_col=8)

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    return out.read()

# --------------------------
# UI
# --------------------------
data_file = st.file_uploader("1) Upload Protection or GI dataset (.csv or .xlsx)", type=["csv", "xlsx"])
template_file = st.file_uploader("2) Upload Provider Excel template (.xlsx)", type=["xlsx"])
product_type = st.selectbox("3) Dataset type", ["Protection", "GI"])

df = None
provider_choice = None
if data_file is not None:
    try:
        df = read_table(data_file)
    except Exception as e:
        st.error(str(e)); st.stop()
    providers = sorted([p for p in df["Provider"].dropna().astype(str).str.strip().unique() if p != ""])
    if not providers:
        st.warning("No providers found in the dataset’s 'Provider' column.")
    provider_choice = st.selectbox("4) Select provider for theming", providers if providers else [""])

ready = (df is not None) and (template_file is not None) and (provider_choice not in [None, "", " "])

if ready:
    if st.button("Build MI Workbook"):
        try:
            xlsx_bytes = build_workbook(template_file.read(), df, provider_choice)
            st.success("Workbook created.")
            st.download_button(
                label="Download MI Workbook",
                data=xlsx_bytes,
                file_name=f"MI_{product_type}_{provider_choice}_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"Failed to build workbook: {e}")
else:
    st.info("Upload the dataset and template, then pick a provider to enable the build button.")

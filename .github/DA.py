import io
import re
import zipfile
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# -----------------------------
# Helpers
# -----------------------------

SUPPORT_REQUIRED_HEADERS = [
    "Date",
    "Source",
    "Description",
    "Reference",
    "Debit (GBP)",
    "Credit (GBP)",
    "Gross (GBP)",
    "VAT (GBP)",
    "Account",
]

MONTHLY_REQUIRED_HEADERS = [
    "DA Firm Name",
    "Sold By",
    "Total Commission Payable",
]

TEMPLATE_HEADERS = [
    "Recruiter",
    "Date Received",
    "Firm",
    "Receipt Name",
    "Payment to TRDA Club",
    "Payment due",
]


def normalise_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalise_header(value):
    return re.sub(r"\s+", " ", normalise_text(value)).lower()


def previous_month_end(today=None):
    today = today or date.today()
    first_day_this_month = today.replace(day=1)
    return first_day_this_month - timedelta(days=1)


def clean_money(series):
    return (
        series.astype(str)
        .str.replace("£", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.strip()
        .replace({"": "0", "nan": "0", "None": "0"})
        .astype(float)
    )


def find_header_row(rows, required_headers):
    required = {normalise_header(h) for h in required_headers}
    for idx, row in enumerate(rows):
        row_headers = {normalise_header(c) for c in row if normalise_text(c)}
        if required.issubset(row_headers):
            return idx
    return None


def normalise_recruiter_name(value):
    name = " ".join(normalise_text(value).split())
    aliases = {
        "steve howard": "Steven Howard",
        "steven howard": "Steven Howard",
    }
    return aliases.get(name.lower(), name)


def parse_support_cash_income(uploaded_file):
    """
    Parses the DA Support Cash Income Report.

    Actual report shape seen in the May 2026 file:
    - One opening company/summary block appears immediately after the headers.
    - That summary block ends with a row whose first cell is "Total".
    - Recruiter sections then repeat as:
        Recruiter name in column A
        transaction rows
        Total row
    - The support report has one header row near the top, not a repeated header per recruiter.
    """
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[cell.value for cell in row] for row in ws.iter_rows()]

    header_idx = find_header_row(rows, SUPPORT_REQUIRED_HEADERS)
    if header_idx is None:
        raise ValueError(
            "Could not find the support report header row containing: "
            + ", ".join(SUPPORT_REQUIRED_HEADERS)
        )

    header_row = rows[header_idx]
    header_map = {}
    for col_idx, value in enumerate(header_row):
        h = normalise_header(value)
        if h:
            header_map[h] = col_idx

    missing = [h for h in SUPPORT_REQUIRED_HEADERS if normalise_header(h) not in header_map]
    if missing:
        raise ValueError(f"Support report is missing columns: {', '.join(missing)}")

    def cell(row, header_name):
        col = header_map[normalise_header(header_name)]
        return row[col] if col < len(row) else None

    parsed_rows = []
    current_recruiter = None
    passed_opening_total = False

    for row_number, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        non_empty = [normalise_text(c) for c in row if normalise_text(c)]
        if not non_empty:
            continue

        first_cell = normalise_text(row[0]) if len(row) > 0 else ""
        first_cell_lower = first_cell.lower()

        # The first block is a company total/summary and should not be assigned to a recruiter.
        # Every recruiter section also ends with a Total row.
        if first_cell_lower.startswith("total"):
            passed_opening_total = True
            current_recruiter = None
            continue

        date_value = cell(row, "Date")
        source_value = cell(row, "Source")
        description_value = cell(row, "Description")
        reference_value = cell(row, "Reference")
        debit_value = cell(row, "Debit (GBP)")
        credit_value = cell(row, "Credit (GBP)")
        gross_value = cell(row, "Gross (GBP)")
        vat_value = cell(row, "VAT (GBP)")
        account_value = cell(row, "Account")

        # Recruiter name rows have just a name in column A and appear only after the opening Total.
        looks_like_recruiter_row = (
            passed_opening_total
            and first_cell
            and len(non_empty) == 1
            and not source_value
            and not description_value
            and not account_value
        )
        if looks_like_recruiter_row:
            current_recruiter = normalise_recruiter_name(first_cell)
            continue

        # A real transaction row has a date and belongs to the current recruiter.
        # Rows before the first recruiter are the opening summary and are intentionally skipped.
        if current_recruiter and date_value:
            parsed_rows.append(
                {
                    "Recruiter": current_recruiter,
                    "Date Received": date_value,
                    "Firm": description_value,
                    "Receipt Name": account_value,
                    "Payment to TRDA Club": credit_value,
                    "Payment due": None,
                    "Source File": "DA Support Cash Income Report",
                    "Debug Row": row_number,
                }
            )

    df = pd.DataFrame(parsed_rows)
    if df.empty:
        raise ValueError(
            "No recruiter transaction rows were found in the support cash income report. "
            "Expected format: opening summary, Total row, then recruiter name rows in column A."
        )
    return df

def parse_monthly_statement(uploaded_file, statement_date):
    """Parses DA-Monthly Statement."""
    raw = pd.read_excel(uploaded_file, sheet_name=0, header=None)
    rows = raw.values.tolist()
    header_idx = find_header_row(rows, MONTHLY_REQUIRED_HEADERS)
    if header_idx is None:
        raise ValueError(
            "Could not find the monthly statement header row containing: "
            + ", ".join(MONTHLY_REQUIRED_HEADERS)
        )

    df = pd.read_excel(uploaded_file, sheet_name=0, header=header_idx)
    df.columns = [normalise_text(c) for c in df.columns]

    missing = [c for c in MONTHLY_REQUIRED_HEADERS if c not in df.columns]
    if missing:
        raise ValueError(f"Monthly statement is missing columns: {', '.join(missing)}")

    df = df[MONTHLY_REQUIRED_HEADERS].copy()
    df = df.dropna(how="all")
    df = df[df["Sold By"].notna()]

    out = pd.DataFrame(
        {
            "Recruiter": df["Sold By"].apply(normalise_recruiter_name),
            "Date Received": statement_date,
            "Firm": df["DA Firm Name"],
            "Receipt Name": "Commission",
            "Payment to TRDA Club": df["Total Commission Payable"],
            "Payment due": None,
            "Source File": "DA-Monthly Statement",
        }
    )
    return out


def write_recruiter_template(template_file, recruiter_df):
    """Writes one recruiter's combined data into a copy of the uploaded template."""
    wb = load_workbook(template_file)
    ws = wb[wb.sheetnames[0]]

    # Ensure headers are in A2:F2.
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        ws.cell(row=2, column=col_idx).value = header

    # Clear old data below row 2 in A:F.
    for row in range(3, ws.max_row + 1):
        for col in range(1, 7):
            ws.cell(row=row, column=col).value = None

    start_row = 3
    for r, (_, row) in enumerate(recruiter_df.iterrows(), start=start_row):
        ws.cell(r, 1).value = row["Recruiter"]
        ws.cell(r, 2).value = row["Date Received"]
        ws.cell(r, 3).value = row["Firm"]
        ws.cell(r, 4).value = row["Receipt Name"]
        ws.cell(r, 5).value = row["Payment to TRDA Club"]
        ws.cell(r, 6).value = row["Payment due"]

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def safe_filename(name):
    cleaned = re.sub(r"[^A-Za-z0-9._ -]+", "", str(name)).strip()
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned or "Unknown_Recruiter"


def build_zip(template_file, combined_df, only_shared):
    zip_buffer = io.BytesIO()

    support_recruiters = set(
        combined_df.loc[combined_df["Source File"] == "DA Support Cash Income Report", "Recruiter"]
        .dropna()
        .astype(str)
        .str.strip()
    )
    monthly_recruiters = set(
        combined_df.loc[combined_df["Source File"] == "DA-Monthly Statement", "Recruiter"]
        .dropna()
        .astype(str)
        .str.strip()
    )

    if only_shared:
        recruiters = sorted(support_recruiters.intersection(monthly_recruiters))
    else:
        recruiters = sorted(set(combined_df["Recruiter"].dropna().astype(str).str.strip()))

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for recruiter in recruiters:
            recruiter_df = combined_df[combined_df["Recruiter"].astype(str).str.strip() == recruiter].copy()
            recruiter_df = recruiter_df[TEMPLATE_HEADERS + ["Source File"]]
            recruiter_df = recruiter_df.sort_values(["Date Received", "Firm"], kind="stable")

            # Need a fresh template stream for each recruiter.
            template_file.seek(0)
            workbook_bytes = write_recruiter_template(template_file, recruiter_df)
            filename = f"{safe_filename(recruiter)}_combined_statement.xlsx"
            zf.writestr(filename, workbook_bytes.getvalue())

    zip_buffer.seek(0)
    return zip_buffer, recruiters, support_recruiters, monthly_recruiters


# -----------------------------
# Streamlit UI
# -----------------------------

st.set_page_config(page_title="DA Statement Builder", page_icon="📄", layout="wide")
st.title("DA Statement Builder")
st.caption("Version v2.3 - fixed support report parser: skips opening Cash Basis block, reads recruiter sections after Total rows")
st.write(
    "Upload the DA Support Cash Income Report, DA-Monthly Statement, and statement template. "
    "The app will create one populated template per recruiter and package them in a ZIP."
)

with st.sidebar:
    st.header("Uploads")
    support_file = st.file_uploader(
        "1. DA Support Cash Income Report",
        type=["xlsx", "xlsm", "xls"],
        key="support_file",
    )
    monthly_file = st.file_uploader(
        "2. DA-Monthly Statement",
        type=["xlsx", "xlsm", "xls"],
        key="monthly_file",
    )
    template_file = st.file_uploader(
        "3. Template",
        type=["xlsx", "xlsm", "xls"],
        key="template_file",
    )

    default_statement_date = previous_month_end()
    statement_date = st.date_input(
        "Date received for monthly statement rows",
        value=default_statement_date,
        help="Defaults to the last day of the previous month.",
    )

    only_shared = st.checkbox(
        "Only create files for recruiters found in both spreadsheets",
        value=True,
    )

run = st.button("Build recruiter statement ZIP", type="primary")

if run:
    if not support_file or not monthly_file or not template_file:
        st.error("Please upload all three files before building the ZIP.")
        st.stop()

    try:
        support_df = parse_support_cash_income(support_file)
        monthly_df = parse_monthly_statement(monthly_file, statement_date)

        combined_df = pd.concat([support_df, monthly_df], ignore_index=True)
        combined_df["Recruiter"] = combined_df["Recruiter"].apply(normalise_recruiter_name)

        zip_buffer, recruiters, support_recruiters, monthly_recruiters = build_zip(
            template_file=template_file,
            combined_df=combined_df,
            only_shared=only_shared,
        )

        st.success(f"Built {len(recruiters)} recruiter statement file(s).")

        col1, col2, col3 = st.columns(3)
        col1.metric("Support report recruiters", len(support_recruiters))
        col2.metric("Monthly statement recruiters", len(monthly_recruiters))
        col3.metric("Files in ZIP", len(recruiters))

        with st.expander("Parser diagnostics", expanded=True):
            st.write("Support report rows by recruiter:")
            st.dataframe(
                support_df["Recruiter"].value_counts().rename_axis("Recruiter").reset_index(name="Rows"),
                use_container_width=True,
            )
            st.write("Monthly statement rows by recruiter:")
            st.dataframe(
                monthly_df["Recruiter"].value_counts().rename_axis("Recruiter").reset_index(name="Rows"),
                use_container_width=True,
            )

        missing_from_monthly = sorted(support_recruiters - monthly_recruiters)
        missing_from_support = sorted(monthly_recruiters - support_recruiters)

        if missing_from_monthly:
            st.warning(
                "Recruiters found in support report but not monthly statement: "
                + ", ".join(missing_from_monthly)
            )
        if missing_from_support:
            st.warning(
                "Recruiters found in monthly statement but not support report: "
                + ", ".join(missing_from_support)
            )

        st.subheader("Preview of combined data")
        st.dataframe(combined_df[TEMPLATE_HEADERS + ["Source File"]], use_container_width=True)

        st.download_button(
            label="Download ZIP",
            data=zip_buffer,
            file_name=f"DA_recruiter_statements_{date.today().isoformat()}.zip",
            mime="application/zip",
        )

    except Exception as exc:
        st.error(str(exc))
        st.exception(exc)

else:
    st.info("Upload the three files, then click Build recruiter statement ZIP.")

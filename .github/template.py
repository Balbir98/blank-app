import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import numbers, Font, Alignment
from datetime import datetime
import zipfile
import os
import time
import tempfile
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

st.title("Commission Statement Generator")

st.markdown("""
Upload the **Zoho Analytics Commission Data** and the **Excel template** to begin.
This tool will generate one statement per AR Firm!
""")

statement_type = st.selectbox("Select Statement Type", ["TRM", "TRB", "TRB - Introducers","Unallocated Cases"])

raw_data_file = st.file_uploader("Upload Commission Data (Excel)", type=["xlsx"], key="raw")
template_file = st.file_uploader("Upload Commission Statement Template (Excel)", type=["xlsx"], key="template")

custom_email_body = ""
if statement_type == "TRM":
    custom_email_body = st.text_area("Optional: Custom Email Body (leave blank to use default)", height=200)

if raw_data_file and template_file:
    if st.button("Run - Generate Statements"):
        start_time = time.time()
        raw_df = pd.read_excel(raw_data_file)

        zip_buffer = io.BytesIO()
        eml_zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf, \
             zipfile.ZipFile(eml_zip_buffer, "w", zipfile.ZIP_DEFLATED) as eml_zip:

            if statement_type == "TRM":
                expected_columns = [
                    "Principal/Adviser Email Address", "AR Firm Name", "Adviser Name", "Date of Statement",
                    "Lender", "Policy Reference", "Product Type", "Client First Name", "Client Surname",
                    "Class", "Commission Payable", "Date Paid to AR"
                ]

                if not all(col in raw_df.columns for col in expected_columns):
                    st.error("The raw data is missing one or more required columns.")
                else:
                    unique_firms = raw_df["AR Firm Name"].dropna().unique()
                    total_firms = len(unique_firms)
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    for i, firm in enumerate(unique_firms):
                        firm_data = raw_df[raw_df["AR Firm Name"] == firm].sort_values(by= "Commission Payable", ascending= False)
                        template_file.seek(0)
                        wb = load_workbook(template_file)
                        ws = wb.active
                        ws["A4"] = firm
                        paid_date = firm_data["Date Paid to AR"].iloc[0] if pd.notnull(firm_data["Date Paid to AR"].iloc[0]) else None
                        ws["H5"] = firm_data["Date Paid to AR"].iloc[0].date() if pd.notnull(firm_data["Date Paid to AR"].iloc[0]) else ""
                        start_row = 7
                        for idx, row in firm_data.iterrows():
                            data_font = Font(name="Calibri", size=9, bold=False)
                            ws.cell(row=start_row, column=1, value=row["Adviser Name"]).font = data_font
                            ws.cell(row=start_row, column=2, value=row["Date of Statement"].date() if pd.notnull(row["Date of Statement"]) else "").font = data_font
                            ws.cell(row=start_row, column=3, value=row["Lender"]).font = data_font
                            ws.cell(row=start_row, column=4, value=row["Policy Reference"]).font = data_font
                            ws.cell(row=start_row, column=5, value=row["Product Type"]).font = data_font
                            ws.cell(row=start_row, column=6, value=row["Client First Name"]).font = data_font
                            ws.cell(row=start_row, column=7, value=row["Client Surname"]).font = data_font
                            ws.cell(row=start_row, column=8, value=row["Class"]).font = data_font
                            commission_cell = ws.cell(row=start_row, column=9, value=row["Commission Payable"])
                            commission_cell.number_format = u"\u00a3#,##0.00"
                            commission_cell.font = data_font
                            start_row += 1

                        output_buffer = io.BytesIO()
                        wb.save(output_buffer)
                        output_buffer.seek(0)
                        formatted_date = paid_date.strftime("%d-%m-%Y") if paid_date else datetime.now().strftime("%d-%m-%Y")
                        filename = f"{firm} {formatted_date}.xlsx"
                        zipf.writestr(filename, output_buffer.getvalue())
                        recipient = firm_data["Principal/Adviser Email Address"].iloc[0]
                        recipient = recipient if pd.notnull(recipient) else ""
                        subject = f"Commission Statement - {firm}"

                        if custom_email_body.strip():
                            formatted_body = custom_email_body.replace('\n', '<br>')
                            html_body = f"""
                                <html>
                                    <body>
                                        <p>{formatted_body}</p>
                                    </body>
                                </html>
                            """

                        else:
                            html_body = f"""
                                <html>
                                    <body>
                                        <p>Good morning,</p>
                                        <p>I hope this email finds you well.</p>
                                        <p>Please find your commission statement for this week attached, payment will be in your account on Friday.</p>
                                        <p><strong>ACRE tips and hints:</strong> <a href='https://rightmortgageadviser.com/commissions-hub/commissions-overview/#getpaidright'>How to get paid right first time!</a></p>
                                        <p>Please ensure:<br>
                                        Mortgages are at <strong>exchanged/complete</strong><br>
                                        Insurances are at <strong>complete</strong><br>
                                        & that <strong>ALL</strong> payments you are expecting, are showing in the accounting tab, one off payments section on the case, in order to show in the payments due section.</p>
                                        <p><a href='https://acresupport.zendesk.com/hc/en-gb/articles/4434905763095-Case-accounting-Adding-one-off-payments-manually-to-a-case'><em>How to add a payment line on acre</em></a></p>
                                        <p>Please visit <a href='https://acresupport.zendesk.com/hc/en-gb/articles/4485360805911-Accounting-view-Introduction'>ACRE ACCOUNTING</a> for information about how you can see what cases have been received and paid, dealing with client fees etc.</p>
                                        <p>For any case queries, please quote the client name, provider/lender and policy number/mortgage account number & we will endeavour to respond to you swiftly.</p>

                                        <br><br>
                                        <p>Kind regards,</p>
                                        <p>
                                            <span style="color: #e36c0a; font-weight: bold;">Commissions Department</span><br>
                                            The Right Mortgage & Protection Network<br>
                                            <span style="color: #e36c0a; font-weight: bold;">TRUST. RESPECT. PARTNERSHIP. OPPORTUNITY</span><br>
                                            <span style="color: #e36c0a; font-weight: bold;">Phone:</span> 01564 732 741<br>
                                            <span style="color: #e36c0a; font-weight: bold;">Web:</span> <a href='https://therightmortgage.co.uk/' target='_blank'>https://therightmortgage.co.uk/</a><br>
                                            70 St Johns Close, Knowle, B93 0NH
                                        </p>
                                        <p style="font-size: 10px; color: black;">
                                            This email and the information it contains may be privileged and/or confidential. It is for the intended addressee(s) only. The unauthorised use, disclosure or copying of this email, or any information it contains is prohibited and could in certain circumstances be a criminal offence. If you are not the intended recipient, please notify <a href='mailto:info@therightmortgage.co.uk'>info@therightmortgage.co.uk</a> immediately and delete the message from your system.<br>
                                            Please note that The Right Mortgage does not enter into any form of contract by means of Internet email. None of the staff of The Right Mortgage is authorised to enter into contracts on behalf of the company in this way. All contracts to which The Right Mortgage is a party are documented by other means.<br>
                                            The Right Mortgage monitors emails to ensure its systems operate effectively and to minimise the risk of viruses. Whilst it has taken reasonable steps to scan this email, it does not accept liability for any virus that it may contain.<br>
                                            Head Office: St Johns Court, 70 St Johns Close, Knowle, Solihull, B93 0NH. Registered in England no. 08130498<br>
                                            The Right Mortgage & Protection Network is a trading style of The Right Mortgage Limited, which is authorised and regulated by the Financial Conduct Authority.
                                        </p>
                                    </body>
                                </html>
                            """


                        msg = MIMEMultipart("mixed")
                        msg["To"] = recipient
                        msg["Subject"] = subject
                        msg.add_header("X-Unsent", "1")
                        
                        alt_part = MIMEMultipart("alternative")
                        alt_part.attach(MIMEText(html_body, "html"))
                        msg.attach(alt_part)
                        from email.mime.application import MIMEApplication
                        part = MIMEApplication(output_buffer.getvalue(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", Name=filename)
                        encoders.encode_base64(part)
                        part.add_header("Content-Disposition", f"attachment; filename=\"{filename}\"")
                        msg.attach(part)
                        eml_filename = f"Email_{firm.replace(' ', '_')}.eml"
                        eml_io = io.BytesIO()
                        from email.generator import BytesGenerator
                        gen = BytesGenerator(eml_io)
                        gen.flatten(msg)
                        eml_io.seek(0)
                        eml_zip.writestr(eml_filename, eml_io.read())
                        elapsed = time.time() - start_time
                        progress_bar.progress((i + 1) / total_firms)
                        status_text.text(f"Processed {i + 1} of {total_firms} firms in {elapsed:.2f} seconds")

            elif statement_type == "TRB":
                column_mapping = {
                    "Adviser": "Adviser",
                    "Date Received": "Date",
                    "Lender": "Lenders",
                    "Policy Number": "Policy reference",
                    "Product Type": "Type",
                    "Client First Name": "First name",
                    "Client Surname": "Surname",
                    "Class": "Class",
                    "Adviser Commission": "Commission"
                }
                if not all(col in raw_df.columns for col in column_mapping.keys()):
                    st.error("The raw TRB data is missing one or more required columns.")
                else:
                    advisers = raw_df["Adviser"].dropna().unique()
                    total_firms = len(advisers)
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    for i, adviser in enumerate(advisers):
                        adviser_data = raw_df[raw_df["Adviser"] == adviser].copy()
                        adviser_data = adviser_data.sort_values(by="Adviser Commission",ascending = False)
                        template_file.seek(0)
                        wb = load_workbook(template_file)
                        ws = wb.active
                        ws["A4"] = adviser
                        paid_date = adviser_data.iloc[0]["Date Paid to Adviser"] if "Date Paid to Adviser" in adviser_data.columns and pd.notnull(adviser_data.iloc[0]["Date Paid to Adviser"]) else None
                        ws["H5"] = adviser_data.iloc[0]["Date Paid to Adviser"].strftime("%d/%m/%Y") if "Date Paid to Adviser" in adviser_data.columns and pd.notnull(adviser_data.iloc[0]["Date Paid to Adviser"]) else ""
                        start_row = 7
                        for idx, row in adviser_data.iterrows():
                            data_font = Font(name="Calibri", size=8, bold=False)
                            for col_index, (src_col, dst_col) in enumerate(column_mapping.items(), start=1):
                                if "Date" in src_col and pd.notnull(row[src_col]):
                                    value = row[src_col].strftime("%d/%m/%Y")
                                else:
                                    value = row[src_col]
                                cell = ws.cell(row=start_row, column=col_index, value=value)
                                cell.font = data_font
                                if dst_col.lower() == "policy reference":
                                    cell.alignment = Alignment(horizontal="left")
                            start_row += 1
                        output_buffer = io.BytesIO()
                        wb.save(output_buffer)
                        output_buffer.seek(0)
                        total_commission = adviser_data['Adviser Commission'].sum()
                        total_str = f"£{total_commission:,.2f}"
                        formatted_date = paid_date.strftime("%d-%m-%Y") if paid_date else datetime.now().strftime("%d-%m-%Y")
                        filename = f"{adviser} {formatted_date} - {total_str}.xlsx"
                        zipf.writestr(filename, output_buffer.getvalue())
                        msg = MIMEMultipart("mixed")
                        recipient = adviser_data["Email"].iloc[0] if "Email" in adviser_data.columns and pd.notnull(adviser_data["Email"].iloc[0]) else ""
                        msg["To"] = recipient
                        msg["Subject"] = f"Commission Statement - {adviser}"
                        msg.add_header("X-Unsent", "1")
                        html_body = f"""
                            <html>
                                <body>
                                    <p>Dear {adviser}! </p>
                                    <p>Attached is your commission statement for this run.</p>
                                    <p><strong>Referred cases</strong><br>
                                    Please check that the amount paid is correct to ensure that amendments don’t need to be deducted at a later date</p>
                                    <p><strong>ACRE tips and hints:</strong><br>
                                    All cases must have an accounting line in order to be paid.<br>
                                    Proc fees must be at ‘exchanged’/‘complete’ to show for payment.<br>
                                    Insurance must be at ‘complete’ to show for payment.</p>
                                    <p>Should you have any queries, please let us know.</p>
                                </body>
                            </html>"""
                        alt_part = MIMEMultipart("alternative")
                        alt_part.attach(MIMEText(html_body, "html"))
                        msg.attach(alt_part)
                        from email.mime.application import MIMEApplication
                        part = MIMEApplication(output_buffer.getvalue(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", Name=filename)
                        encoders.encode_base64(part)
                        part.add_header("Content-Disposition", f"attachment; filename=\"{filename}\"")
                        msg.attach(part)
                        eml_filename = f"Email_TRB_{adviser.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}_{total_str}.eml"
                        eml_io = io.BytesIO()
                        from email.generator import BytesGenerator
                        gen = BytesGenerator(eml_io)
                        gen.flatten(msg)
                        eml_io.seek(0)
                        eml_zip.writestr(eml_filename, eml_io.read())
                        elapsed = time.time() - start_time
                        progress_bar.progress((i + 1) / total_firms)
                        status_text.text(f"Processed {i + 1} of {total_firms} advisers in {elapsed:.2f} seconds")
            elif statement_type == "TRB - Introducers":
                            column_mapping = {
                                "Adviser": "Adviser",
                                "Date Received": "Date",
                                "Lender": "Lenders",
                                "Policy Number": "Policy reference",
                                "Product Type": "Type",
                                "Client First Name": "First name",
                                "Client Surname": "Surname",
                                "Class": "Class",
                                "Introducer Commission": "Commission"
                            }
                            if not all(col in raw_df.columns for col in column_mapping.keys()):
                                st.error("The raw TRB Introducers data is missing one or more required columns.")
                            else:
                                introducers = raw_df["Introducer"].dropna().unique()
                                total_firms = len(introducers)
                                progress_bar = st.progress(0)
                                status_text = st.empty()
                                for i, introducer in enumerate(introducers):
                                    introducer_data = raw_df[raw_df["Introducer"] == introducer].copy()
                                    introducer_data = introducer_data.sort_values(by="Introducer Commission",ascending = False)
                                    template_file.seek(0)
                                    wb = load_workbook(template_file)
                                    ws = wb.active
                                    ws["A4"] = introducer
                                    paid_date = introducer_data.iloc[0]["Date Paid to Introducer"] if "Date Paid to Introducer" in introducer_data.columns and pd.notnull(introducer_data.iloc[0]["Date Paid to Introducer"]) else None
                                    ws["H5"] = introducer_data.iloc[0]["Date Paid to Introducer"].strftime("%d/%m/%Y") if "Date Paid to Introducer" in introducer_data.columns and pd.notnull(introducer_data.iloc[0]["Date Paid to Introducer"]) else ""
                                    start_row = 7
                                    for idx, row in introducer_data.iterrows():
                                        data_font = Font(name="Calibri", size=8, bold=False)
                                        for col_index, (src_col, dst_col) in enumerate(column_mapping.items(), start=1):
                                            if "Date" in src_col and pd.notnull(row[src_col]):
                                                value = row[src_col].strftime("%d/%m/%Y")
                                            else:
                                                value = row[src_col]
                                            cell = ws.cell(row=start_row, column=col_index, value=value)
                                            cell.font = data_font
                                            if dst_col.lower() == "policy reference":
                                                cell.alignment = Alignment(horizontal="left")
                                        start_row += 1
                                    output_buffer = io.BytesIO()
                                    wb.save(output_buffer)
                                    output_buffer.seek(0)
                                    total_commission = introducer_data['Introducer Commission'].sum()
                                    total_str = f"£{total_commission:,.2f}"
                                    formatted_date = paid_date.strftime("%d-%m-%Y") if paid_date else datetime.now().strftime("%d-%m-%Y")
                                    filename = f"{introducer} {formatted_date} - {total_str}.xlsx"
                                    zipf.writestr(filename, output_buffer.getvalue())
                                    msg = MIMEMultipart("mixed")
                                    recipient = introducer_data["Introducer Email"].iloc[0] if "Introducer Email" in introducer_data.columns and pd.notnull(introducer_data["Introducer Email"].iloc[0]) else ""
                                    msg["To"] = recipient
                                    msg["Subject"] = f"Commission Statement - {introducer}"
                                    msg.add_header("X-Unsent", "1")
                                    html_body = f"""
                                        <html>
                                            <body>
                                                <p>Attached is your commission statement for this run.</p>
                                                <p><strong>Referred cases</strong><br>
                                                Please check that the amount paid is correct to ensure that amendments don’t need to be deducted at a later date</p>
                                                <p><strong>ACRE tips and hints:</strong><br>
                                                All cases must have an accounting line in order to be paid.<br>
                                                Proc fees must be at ‘exchanged’/‘complete’ to show for payment.<br>
                                                Insurance must be at ‘complete’ to show for payment.</p>
                                                <p>Should you have any queries, please let us know.</p>
                                            </body>
                                        </html>"""
                                    alt_part = MIMEMultipart("alternative")
                                    alt_part.attach(MIMEText(html_body, "html"))
                                    msg.attach(alt_part)
                                    from email.mime.application import MIMEApplication
                                    part = MIMEApplication(output_buffer.getvalue(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", Name=filename)
                                    encoders.encode_base64(part)
                                    part.add_header("Content-Disposition", f"attachment; filename=\"{filename}\"")
                                    msg.attach(part)
                                    eml_filename = f"Email_TRB_Introducer_{introducer.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}_{total_str}.eml"
                                    eml_io = io.BytesIO()
                                    from email.generator import BytesGenerator
                                    gen = BytesGenerator(eml_io)
                                    gen.flatten(msg)
                                    eml_io.seek(0)
                                    eml_zip.writestr(eml_filename, eml_io.read())
                                    elapsed = time.time() - start_time
                                    progress_bar.progress((i + 1) / total_firms)
                                    status_text.text(f"Processed {i + 1} of {total_firms} introducers in {elapsed:.2f} seconds")
            elif statement_type == "Unallocated Cases":
                expected_columns = [
                    "AR Firm Name", "Email", "Adviser Name", "Lenders", "Policy Reference",
                    "Product Type", "Client First Name", "Client Surname", "Class"
                ]

                if not all(col in raw_df.columns for col in expected_columns):
                    st.error("The raw Unallocated data is missing one or more required columns.")
                else:
                    firms = raw_df["AR Firm Name"].dropna().unique()
                    total_firms = len(firms)
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    for i, firm in enumerate(firms):
                        firm_data = raw_df[raw_df["AR Firm Name"] == firm].copy()
                        firm_data = firm_data.sort_values(by="Adviser Name")

                        template_file.seek(0)
                        wb = load_workbook(template_file)
                        ws = wb.active

                        ws["A4"] = firm
                        ws["F4"] = datetime.now().strftime("%d/%m/%Y")

                        start_row = 7
                        for idx, row in firm_data.iterrows():
                            data_font = Font(name="Calibri", size=9, bold=False)
                            ws.cell(row=start_row, column=1, value=row["Adviser Name"]).font = data_font
                            ws.cell(row=start_row, column=2, value=row["Lenders"]).font = data_font
                            ws.cell(row=start_row, column=3, value=row["Policy Reference"]).font = data_font
                            ws.cell(row=start_row, column=4, value=row["Product Type"]).font = data_font
                            ws.cell(row=start_row, column=5, value=row["Client First Name"]).font = data_font
                            ws.cell(row=start_row, column=6, value=row["Client Surname"]).font = data_font
                            ws.cell(row=start_row, column=7, value=row["Class"]).font = data_font
                            start_row += 1

                        output_buffer = io.BytesIO()
                        wb.save(output_buffer)
                        output_buffer.seek(0)

                        formatted_date = datetime.now().strftime("%d-%m-%Y")
                        filename = f"Unallocated - {firm} - {formatted_date}.xlsx"

                        # Write Excel to ZIP archive directly (flat)
                        zipf.writestr(filename, output_buffer.getvalue())

                        recipient = firm_data["Email"].iloc[0] if pd.notnull(firm_data["Email"].iloc[0]) else ""
                        subject = f"Unallocated Report - {firm}"

                        html_body = f"""
                            <html>
                                <body>
                                    <p><strong>Classification - <span style='color: red;'>Confidential</span></strong></p>
                                    <p>Good morning,</p>
                                    <p>I hope this email finds you well.</p>
                                    <p>Please find your unallocated commission breakdown.<br>
                                    If you can let us know when the cases have been completed on acre.</p>
                                    <p>Any commission statement due this week will be sent separately</p>
                                    <p>For any case queries, please quote the client name, provider/lender and policy number/mortgage account number & we will endeavour to respond to you swiftly.</p>
                                    <br>
                                    <p style="font-family: Calibri, sans-serif; font-size: 15px;">
                                        <strong style="color:#e57200">Commissions Department</strong><br>
                                        The Right Mortgage & Protection Network<br>
                                        <strong style="color:#e57200">TRUST. RESPECT. PARTNERSHIP. OPPORTUNITY</strong><br>
                                        <strong style="color:#e57200">Phone:</strong> 01564 732 741<br>
                                        <strong style="color:#e57200">Web:</strong> <a href="https://therightmortgage.co.uk/">https://therightmortgage.co.uk/</a><br>
                                        70 St Johns Close, Knowle, B93 0NH
                                    </p>
                                    <p style="font-size:12px; color:#000;">
                                    This email and the information it contains may be privileged and/or confidential. It is for the intended addressee(s) only. The unauthorised use, disclosure or copying of this email, or any information it contains is prohibited and could in certain circumstances be a criminal offence. If you are not the intended recipient, please notify <a href='mailto:info@therightmortgage.co.uk'>info@therightmortgage.co.uk</a> immediately and delete the message from your system.<br>
                                    Please note that The Right Mortgage does not enter into any form of contract by means of Internet email. None of the staff of The Right Mortgage is authorised to enter into contracts on behalf of the company in this way. All contracts to which The Right Mortgage is a party are documented by other means.<br>
                                    The Right Mortgage monitors emails to ensure its systems operate effectively and to minimise the risk of viruses. Whilst it has taken reasonable steps to scan this email, it does not accept liability for any virus that it may contain.<br>
                                    Head Office: St Johns Court, 70 St Johns Close, Knowle, Solihull, B93 0NH. Registered in England no. 08130498<br>
                                    The Right Mortgage & Protection Network is a trading style of The Right Mortgage Limited, which is authorised and regulated by the Financial Conduct Authority.
                                    </p>
                                </body>
                            </html>
                        """

                        msg = MIMEMultipart("mixed")
                        msg["To"] = recipient
                        msg["Subject"] = subject
                        msg.add_header("X-Unsent", "1")

                        alt_part = MIMEMultipart("alternative")
                        alt_part.attach(MIMEText(html_body, "html"))
                        msg.attach(alt_part)

                        from email.mime.application import MIMEApplication
                        part = MIMEApplication(output_buffer.getvalue(), _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", Name=filename)
                        encoders.encode_base64(part)
                        part.add_header("Content-Disposition", f"attachment; filename=\"{filename}\"")
                        msg.attach(part)

                        eml_filename = f"Unallocated - {firm.replace('/', '-') } - {formatted_date}.eml"
                        eml_io = io.BytesIO()
                        from email.generator import BytesGenerator
                        gen = BytesGenerator(eml_io)
                        gen.flatten(msg)
                        eml_io.seek(0)
                        eml_zip.writestr(eml_filename, eml_io.read())

                        elapsed = time.time() - start_time
                        progress_bar.progress((i + 1) / total_firms)
                        status_text.text(f"Processed {i + 1} of {total_firms} firms in {elapsed:.2f} seconds")


        zip_buffer.seek(0)
        eml_zip_buffer.seek(0)

        st.download_button(
            label="Download All Statements as ZIP",
            data=zip_buffer,
            file_name="All_Commission_Statements.zip",
            mime="application/zip"
        )

        st.download_button(
            label="Download Draft Emails as EML ZIP",
            data=eml_zip_buffer,
            file_name="All_Email_Drafts.zip",
            mime="application/zip"
        )

        total_time = time.time() - start_time
        st.success(f"All statements and email drafts generated successfully in {total_time:.2f} seconds!")

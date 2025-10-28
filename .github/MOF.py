import re
import os
from io import BytesIO
import zipfile
import pandas as pd
import streamlit as st

# Excel handling / styling
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

# ===========================
# Utilities: robust readers
# ===========================
def _read_any_table(uploaded_file, preferred_sheet_name=None):
    name = uploaded_file.name.lower()
    ext = os.path.splitext(name)[1]
    if ext in [".csv", ".txt"]:
        try:
            return pd.read_csv(uploaded_file, engine="python", sep=None)
        except Exception:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, engine="python", sep=None, encoding="latin-1")
    elif ext in [".xlsx", ".xls"]:
        if preferred_sheet_name:
            try:
                return pd.read_excel(uploaded_file, sheet_name=preferred_sheet_name)
            except Exception:
                uploaded_file.seek(0)
        return pd.read_excel(uploaded_file, sheet_name=0)
    else:
        try:
            return pd.read_csv(uploaded_file, engine="python", sep=None)
        except Exception:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, engine="python", sep=None, encoding="latin-1")

# ===========================
# Type -> Product overrides
# ===========================
TYPE_TO_PRODUCT = {
    "Compliance Update Sponsorship": "Monthly Compliance Bulletin",
    "Podcast": "Podcasts",
    "Directly Authorised Club Site Takeover": "Directly Authorised Club Site Takeover",
    "Full Takeover (DA Club and Adviser Site)": "Full Takeover (DA Club and Adviser Site)",
    "Network Adviser Site Takeover": "Network Adviser Site Takeover",
    "Promotional Emails": "Promotional Email",
    "Social Media Post Share": "Social Media Post",
    "Training Video": "Online training video (provider produces and edits)",
    "Video adverts": "Video advert",
    "Product Focus emails": "Product Focus Emails",
    # explicit 1:1 requests
    "Compliance Webinar Sponsorship": "Compliance Webinar Sponsorship",
    "Full Adviser Site Takeover (Network and DA Club)": "Full Adviser Site Takeover (Network and DA Club)",
    "Full Adviser Site Login Takeover (Network and DA Club)": "Full Adviser Site Login Takeover (Network and DA Club)",
    "Network Adviser Site Login Takeover": "Network Adviser Site Login Takeover",
}
_TYPE_OVERRIDE_LC = {k.casefold().strip(): v for k, v in TYPE_TO_PRODUCT.items()}

def _apply_type_overrides(df: pd.DataFrame) -> pd.DataFrame:
    if 'Type' not in df.columns or 'Product' not in df.columns:
        return df
    def _override(row):
        t = str(row.get('Type', '')).casefold().strip()
        return _TYPE_OVERRIDE_LC.get(t, row.get('Product'))
    df = df.copy()
    df['Product'] = df.apply(_override, axis=1)
    return df

# ===========================
# Product aliasing (canonical names for MOF match)
# ===========================
PRODUCT_ALIASES = {
    "product focus emails": "Product Focus Emails",
    "product focus email": "Product Focus Emails",
    "promotional emails": "Promotional Emails",
    "promotional email": "Promotional Emails",
    "social media post share": "Social Media Post Share",
    "social media post": "Social Media Post Share",
    "compliance update sponsorship email": "Compliance Update Sponsorship Email",
    "compliance update sponsorship": "Compliance Update Sponsorship Email",
}
def _apply_product_aliases(s):
    if pd.isna(s): return s
    v = str(s).strip()
    return PRODUCT_ALIASES.get(v.casefold(), v)

# ===========================
# Helper: loose column picker
# ===========================
def _pick(colnames, *candidates):
    lowmap = {str(c).strip().lower(): c for c in colnames}
    for cand in candidates:
        if cand is None:
            continue
        key = str(cand).strip().lower()
        if key in lowmap:
            return lowmap[key]
    lc = [str(c).lower() for c in colnames]
    for cand in candidates:
        if not cand:
            continue
        ck = str(cand).strip().lower()
        for i, name in enumerate(lc):
            if ck == name:
                return colnames[i]
    return None

# ===========================
# Wishlist Transformation
# ===========================
REPEATED_FIRST = [
    'Random ID','Provider Name','Name','Phone','Email',
    'Events Name','Events Email',
    'Marketing Publications Name','Marketing Publications Email',
    'Copy Name','Copy Email',
    'When To Invoice',
    'Invoice Name','Invoice Email',
]

# Two main notes questions (shown in Notes tab only)
MAIN_NOTES_QUESTIONS = [
    "Please provide any feedback on our Marketing & Opportunities 2026 Pack and webinar:",
    "Please provide any further notes you may have or want to have considered with this form:",
]
NOTES_SOURCE_HEADERS = MAIN_NOTES_QUESTIONS + ["Further notes", "Any further notes"]

# Admin/ID fields (NOT notes)
ID_COLS_WISHLIST = REPEATED_FIRST.copy()

def _is_event_label(x):
    if pd.isna(x): return False
    s = str(x).strip()
    months3 = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']
    if any(s.lower().startswith(m) for m in months3): return True
    if s in ['Q1','Q2','Q3','Q4','Monthly','Quarterly']: return True
    if re.search(r'\d', s): return True
    return False

# ===== Regional Roadshow helpers =====
_LOCATION_SET = {
    "north","south","east","west","midlands","central","scotland","wales","wales/bristol","wales / bristol",
    "northern ireland","north east","north west","south east","south west","london",
    "yorkshire","humberside","solihull","bristol"
}
def _looks_like_location(s: str) -> bool:
    if not isinstance(s, str): s = str(s or "")
    return s.strip().casefold() in _LOCATION_SET

def _is_rre_type(t: str) -> bool:
    if not isinstance(t, str): t = str(t or "")
    return t.casefold().strip().startswith("regional roadshow event")

def _is_email_marketing_type(t: str) -> bool:
    if not isinstance(t, str): t = str(t or "")
    return "email marketing" in t.casefold()

def _is_unchecked(v) -> bool:
    if pd.isna(v): return True
    s = str(v).strip().casefold()
    return s in {"", "no", "false", "0", "none", "n/a", "na", "not selected", "unchecked"}

# Types where a single cell may contain multiple comma-separated product choices
SPLIT_ON_COMMA_TYPES = {
    "web page on adviser site",
    "equity release workshops",
}

def transform_wishlist(form_df: pd.DataFrame, costs_df: pd.DataFrame) -> pd.DataFrame:
    """
    Parse Zoho wide export (row 0 = subheaders) → long rows, then join Cost by (Type, Product).
    """
    if form_df.shape[0] < 2:
        return pd.DataFrame(columns=REPEATED_FIRST + ['Type','Event Date (if applicable)','Product','Cost','F2F or Online?'])

    subheaders = form_df.iloc[0]
    data_rows = form_df.iloc[1:].reset_index(drop=True)

    col_lc_map = {str(c).strip().lower(): c for c in form_df.columns}
    wanted_cols = [col_lc_map.get(c.lower()) for c in ID_COLS_WISHLIST if col_lc_map.get(c.lower())]

    # Detect notes columns present
    notes_cols = [col_lc_map[h.lower()] for h in NOTES_SOURCE_HEADERS if h.lower() in col_lc_map]
    notes_name_set = set(notes_cols)
    def _is_notes_column(col_name) -> bool:
        return col_name in notes_name_set

    # Resolve the two *main* notes columns if present
    q1_col = col_lc_map.get(MAIN_NOTES_QUESTIONS[0].lower())
    q2_col = col_lc_map.get(MAIN_NOTES_QUESTIONS[1].lower())

    # ---- Parse into line items (excluding notes columns) ----
    records = []
    for ridx, row in data_rows.iterrows():
        current_type = None
        for j, col in enumerate(form_df.columns):
            if not str(col).startswith('Unnamed'):
                if col not in wanted_cols and col not in ['Added Time','Referrer Name','Task Owner'] and not _is_notes_column(col):
                    current_type = col

            val = row[col]
            if pd.isna(val):
                continue
            text_val = str(val).strip()
            sub = subheaders.iloc[j] if j < len(subheaders) else None

            if _is_notes_column(col):
                continue

            # Email Marketing: month + comma-separated selections → 1 row per selection
            if current_type and _is_email_marketing_type(current_type) and sub is not None and _is_event_label(sub):
                if _is_unchecked(text_val):
                    continue
                parts = [p.strip() for p in re.split(r'\s*,\s*', text_val) if p.strip()]
                if parts:
                    for p in parts:
                        records.append({
                            '_ridx': ridx,
                            'Type': current_type,
                            'Event Date (if applicable)': str(sub).strip(),
                            'Product': p
                        })
                    continue

            # Regional roadshow matrix: Location (subheader) → Event Date, cell text → Product
            if current_type and _is_rre_type(current_type) and sub is not None and _looks_like_location(str(sub)):
                if _is_unchecked(text_val):
                    continue
                records.append({
                    '_ridx': ridx,
                    'Type': current_type,
                    'Event Date (if applicable)': str(sub).strip(),
                    'Product': text_val
                })
                continue

            # Default behaviour (plus comma-split for specific Types)
            if not str(col).startswith('Unnamed'):
                if col in wanted_cols or col in ['Added Time','Referrer Name','Task Owner']:
                    continue
                if re.search(r'\boption(s)?\b', text_val, flags=re.I) and not pd.isna(sub):
                    prod = str(sub).strip()
                    evt = None if not _is_event_label(sub) else str(sub).strip()
                    if _is_event_label(sub):
                        prod = text_val
                else:
                    prod = text_val
                    evt = str(sub).strip() if _is_event_label(sub) else None
                # Split-on-comma for designated Types
                if current_type and current_type.casefold().strip() in SPLIT_ON_COMMA_TYPES and ',' in str(prod):
                    for p in [x.strip() for x in str(prod).split(',') if x.strip()]:
                        records.append({'_ridx': ridx, 'Type': current_type, 'Event Date (if applicable)': evt, 'Product': p})
                else:
                    records.append({'_ridx': ridx, 'Type': current_type, 'Event Date (if applicable)': evt, 'Product': prod})
            else:
                if re.search(r'\boption(s)?\b', text_val, flags=re.I):
                    prod = str(sub).strip() if not pd.isna(sub) else text_val
                    evt = None
                else:
                    if _is_event_label(sub):
                        evt = str(sub).strip()
                        prod = text_val
                    else:
                        evt = None
                        prod = str(sub).strip() if not pd.isna(sub) else text_val
                # Split-on-comma for designated Types
                if current_type and current_type.casefold().strip() in SPLIT_ON_COMMA_TYPES and ',' in str(prod):
                    for p in [x.strip() for x in str(prod).split(',') if x.strip()]:
                        records.append({'_ridx': ridx, 'Type': current_type, 'Event Date (if applicable)': evt, 'Product': p})
                else:
                    if current_type is None:
                        continue
                    records.append({'_ridx': ridx, 'Type': current_type, 'Event Date (if applicable)': evt, 'Product': prod})

    out = pd.DataFrame.from_records(records)

    # Attach repeated fields
    for c in REPEATED_FIRST:
        if c in data_rows.columns:
            out[c] = out['_ridx'].map(data_rows[c])
        else:
            out[c] = None

    # Attach per-provider notes answers (hidden)
    if q1_col:
        q1_by_ridx = data_rows[q1_col]
        out['_note_q1'] = out['_ridx'].map(q1_by_ridx).astype(str).apply(lambda s: "" if s.lower() == "nan" else s).fillna("")
    else:
        out['_note_q1'] = ""
    if q2_col:
        q2_by_ridx = data_rows[q2_col]
        out['_note_q2'] = out['_ridx'].map(q2_by_ridx).astype(str).apply(lambda s: "" if s.lower() == "nan" else s).fillna("")
    else:
        out['_note_q2'] = ""

    if out.empty:
        return pd.DataFrame(columns=REPEATED_FIRST + ['Type','Event Date (if applicable)','Product','Cost','F2F or Online?'])

    # Remove any stray rows where Type equals a notes header (safety)
    notes_lc = {h.casefold().strip() for h in NOTES_SOURCE_HEADERS}
    out = out[~out['Type'].astype(str).str.casefold().str.strip().isin(notes_lc)].copy()

    # Keep key outputs (plus hidden notes)
    out = out[['_ridx'] + REPEATED_FIRST + ['Type','Event Date (if applicable)','Product','_note_q1','_note_q2']].copy()

    # Safety: if a regional row accidentally has the location in Product, blank it
    mask_rre = out['Type'].apply(_is_rre_type).fillna(False)
    out.loc[mask_rre & out['Product'].apply(_looks_like_location), 'Product'] = None

    # Apply product aliasing then type overrides
    out['Product'] = out['Product'].apply(_apply_product_aliases)
    out = _apply_type_overrides(out)

    # ---- Costs join by (Type, Product) ----
    def _norm(s): return None if pd.isna(s) else str(s).strip()
    if not {'Product','Cost','Type'}.issubset(costs_df.columns):
        raise ValueError("Cost sheet must contain columns: Type, Product, Cost")

    costs2 = costs_df.copy()
    costs2['Type_norm'] = costs2['Type'].apply(_norm)
    costs2['Product_norm'] = costs2['Product'].apply(_norm).apply(_apply_product_aliases)

    bring_cols = ['Type_norm','Product_norm','Cost']
    f2f_col_name = None
    for cand in ['F2F or Online?', 'F2F or Online', 'F2F/Online']:
        if cand in costs2.columns:
            bring_cols.append(cand)
            f2f_col_name = cand
            break
    costs2 = costs2[bring_cols].drop_duplicates(subset=['Type_norm','Product_norm'], keep='first')

    out['Type_norm'] = out['Type'].apply(_norm)
    out['Product_norm'] = out['Product'].apply(_norm).apply(_apply_product_aliases)

    out = out.merge(costs2, on=['Type_norm','Product_norm'], how='left').drop(columns=['Type_norm','Product_norm'])

    # Final internal columns (include hidden notes answers)
    final_cols_internal = REPEATED_FIRST + ['Type','Event Date (if applicable)','Product','Cost']
    if f2f_col_name:
        out = out.rename(columns={f2f_col_name: 'F2F or Online?'})
        final_cols_internal += ['F2F or Online?']
    final_cols_internal += ['_note_q1','_note_q2']
    out = out[final_cols_internal].sort_values(['Random ID','Provider Name','Name','Type','Event Date (if applicable)','Product']).reset_index(drop=True)

    return out

# ===========================
# Template helpers (shared)
# ===========================
def _sanitize_name(name: str) -> str:
    if pd.isna(name): return ""
    s = str(name).replace(",", " ")
    return re.sub(r"\s+", " ", s).strip()

def _find_table_header_row(ws):
    max_row = min(ws.max_row, 200)
    max_col = min(ws.max_column, 80)
    for r in range(1, max_row + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        for c in range(1, max_col):
            v = row_vals[c-1]
            if v is not None and str(v).strip().lower() == "type":
                nxt = row_vals[c] if c < max_col else None
                if nxt is not None and str(nxt).strip().lower() == "product":
                    hmap = {}
                    for c2 in range(1, max_col + 1):
                        v2 = ws.cell(r, c2).value
                        if v2 is not None:
                            hmap[str(v2).strip()] = c2
                    return r, c, hmap
    raise RuntimeError("Could not find the table header row with 'Type' and 'Product'.")

def _get_col(hmap, key, aliases=()):
    if key in hmap: return hmap[key]
    lowered = {k.lower(): v for k, v in hmap.items()}
    if key.lower() in lowered: return lowered[key.lower()]
    for a in aliases:
        if a in hmap: return hmap[a]
        if a.lower() in lowered: return lowered[a.lower()]
    return None

def _canon_label(s: str) -> str:
    return re.sub(r'[^a-z0-9]+', '', str(s).lower())

def _find_label_in_column(ws, label_text, col_idx, row_start, row_end):
    tgt = _canon_label(label_text)
    row_end = min(row_end, ws.max_row)
    for rr in range(row_start, row_end + 1):
        v = ws.cell(rr, col_idx).value
        if v is None:
            continue
        if tgt in _canon_label(v):
            return rr, col_idx
    return None, None

def _first_value_cell_right(ws, r, c, try_two=True):
    for cc in [c + 1, c + 2] if try_two else [c + 1]:
        v = ws.cell(r, cc).value
        if isinstance(v, str) and v.strip() == "£":
            continue
        return ws.cell(r, cc)
    return ws.cell(r, c + 1)

# Notes sheet helper (find or create)
def _get_notes_ws(wb):
    for name in wb.sheetnames:
        if str(name).strip().casefold() == "notes":
            return wb[name]
    return wb.create_sheet("Notes")

# ===========================
# Template population
# ===========================
def _populate_template_bytes(template_bytes: bytes, cleaned: pd.DataFrame, costs_df: pd.DataFrame | None) -> BytesIO:
    # Build a (Type, Product)->F2F map if we have a cost sheet
    f2f_map = {}
    if costs_df is not None:
        def _n(s): return None if pd.isna(s) else str(s).strip()
        tmp = costs_df.copy()
        if 'F2F or Online?' in tmp.columns:
            tmp['Type_norm'] = tmp['Type'].apply(_n)
            tmp['Product_norm'] = tmp['Product'].apply(_n).apply(_apply_product_aliases)
            f2f_map = dict(zip(zip(tmp['Type_norm'], tmp['Product_norm']), tmp['F2F or Online?']))

    zip_buf = BytesIO()
    dotted = Side(style='dotted')
    header_font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")  # white header text
    ACC_FMT = '_-£* #,##0.00_-;_-£* -#,##0.00_-;_-£* "-"??_-;_-@_-'

    with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        # GROUP BY provider + contact (separate template per submitter)
        for (provider, contact_name, contact_email), dfp in cleaned.groupby(['Provider Name','Name','Email'], dropna=False):
            wb = load_workbook(BytesIO(template_bytes))
            ws = wb.active

            # Fixed cells
            prov = "" if pd.isna(provider) else str(provider)
            nm   = _sanitize_name(contact_name)
            ph   = "" if 'Phone' not in dfp.columns else ("" if pd.isna(dfp['Phone'].iloc[0]) else str(dfp['Phone'].iloc[0]))
            em   = "" if pd.isna(contact_email) else str(contact_email)
            wti  = dfp['When To Invoice'].iloc[0] if 'When To Invoice' in dfp.columns else ""

            ws['B4'] = prov
            ws['B6'] = nm
            ws['D4'] = wti
            ws['D6'] = ph
            ws['F6'] = em   # email goes here

            # Secondary contacts
            def _first(dfcol):
                return (dfcol.iloc[0] if dfcol is not None and len(dfp) else "")
            if 'Events Name' in dfp.columns: ws['B10'] = _first(dfp['Events Name'])
            if 'Events Email' in dfp.columns: ws['B12'] = _first(dfp['Events Email'])
            if 'Marketing Publications Name' in dfp.columns: ws['D10'] = _first(dfp['Marketing Publications Name'])
            if 'Marketing Publications Email' in dfp.columns: ws['D12'] = _first(dfp['Marketing Publications Email'])
            if 'Invoice Name' in dfp.columns: ws['F10'] = _first(dfp['Invoice Name'])
            if 'Invoice Email' in dfp.columns: ws['F12'] = _first(dfp['Invoice Email'])
            if 'Copy Name' in dfp.columns: ws['H10'] = _first(dfp['Copy Name'])
            if 'Copy Email' in dfp.columns: ws['H12'] = _first(dfp['Copy Email'])

            # Find table header
            hdr_row, hdr_col, hmap = _find_table_header_row(ws)
            c_Type   = _get_col(hmap, "Type")
            c_Prod   = _get_col(hmap, "Product")
            c_Det    = _get_col(hmap, "Details")
            c_Date   = _get_col(hmap, "Date")
            c_Charge = _get_col(hmap, "Charge", aliases=("Cost","Price"))
            c_Qty    = _get_col(hmap, "Qty", aliases=("Quantity",))
            c_Total  = _get_col(hmap, "Total")
            c_Notes  = _get_col(hmap, "Notes")

            # Header font = white
            for c in [c_Type,c_Prod,c_Det,c_Date,c_Charge,c_Qty,c_Total,c_Notes]:
                if c: ws.cell(hdr_row, c).font = header_font

            # Insert rows for n items
            start_row = hdr_row + 1
            n = len(dfp)
            if n > 1:
                ws.insert_rows(start_row + 1, amount=n - 1)

            # Fill rows
            for i, (_, r) in enumerate(dfp.iterrows()):
                rr = start_row + i
                typ   = r.get("Type", "")
                prod  = r.get("Product", "")
                datev = r.get("Event Date (if applicable)", "")
                charge = r.get("Cost", None)
                qty = 1
                details = r.get("F2F or Online?", "")

                if not details and costs_df is not None:
                    k = (str(typ).strip(), _apply_product_aliases(str(prod).strip()))
                    details = f2f_map.get(k, "")

                if c_Type:   ws.cell(rr, c_Type, typ)
                if c_Prod:   ws.cell(rr, c_Prod, prod)
                if c_Det:    ws.cell(rr, c_Det, details)
                if c_Date:   ws.cell(rr, c_Date, datev)
                if c_Qty:    ws.cell(rr, c_Qty, qty)
                if c_Charge:
                    ws.cell(rr, c_Charge, charge).number_format = ACC_FMT
                if c_Total:
                    try:
                        total_val = (qty or 0) * (float(charge) if charge not in [None, ""] else 0.0)
                    except Exception:
                        total_val = None
                    ws.cell(rr, c_Total, total_val).number_format = ACC_FMT
                if c_Notes:
                    ws.cell(rr, c_Notes, None)

            # Borders across table (dotted)
            last_row = start_row + max(n - 1, 0)
            table_cols = [c for c in [c_Type,c_Prod,c_Det,c_Date,c_Charge,c_Qty,c_Total,c_Notes] if c]
            first_col, last_col = min(table_cols), max(table_cols)
            dotted_b = Border(top=dotted, bottom=dotted, left=dotted, right=dotted)
            for r in range(hdr_row, last_row + 1):
                for c in range(first_col, last_col + 1):
                    ws.cell(r, c).border = dotted_b

            # ---------- Summary block ----------
            ACC = ACC_FMT
            sum_col = c_Total if c_Total else c_Charge
            if sum_col:
                sum_rng = f"{get_column_letter(sum_col)}{start_row}:{get_column_letter(sum_col)}{last_row}"

                # Detect label column from "Total Package"
                label_col = None
                for c in range(1, ws.max_column + 1):
                    for rr in range(last_row + 1, min(ws.max_row, last_row + 200) + 1):
                        v = ws.cell(rr, c).value
                        if v and _canon_label(v) == _canon_label("Total Package"):
                            label_col = c
                            break
                    if label_col:
                        break
                if not label_col:
                    label_col = 8

                value_col = label_col + 1
                search_start = last_row + 1
                search_end   = min(ws.max_row, last_row + 200)

                r_tp, _ = _find_label_in_column(ws, "Total Package", label_col, search_start, search_end)
                tp_coord = None
                if r_tp:
                    tp_cell = _first_value_cell_right(ws, r_tp, label_col)
                    tp_cell.value = f"=SUM({sum_rng})"
                    tp_cell.number_format = ACC
                    tp_coord = tp_cell.coordinate

                r_disc, _ = _find_label_in_column(ws, "Discount", label_col, search_start, search_end)
                disc_coord = None
                if r_disc:
                    disc_cell = _first_value_cell_right(ws, r_disc, label_col)
                    if disc_cell.value is None:
                        disc_cell.value = 0
                    disc_cell.number_format = ACC
                    disc_coord = disc_cell.coordinate

                r_tpp, _ = _find_label_in_column(ws, "Total Package Price", label_col, search_start, search_end)
                tpp_coord = None
                if r_tpp and tp_coord and disc_coord:
                    tpp_cell = _first_value_cell_right(ws, r_tpp, label_col)
                    tpp_cell.value = f"={tp_coord}-{disc_coord}"
                    tpp_cell.number_format = ACC
                    tpp_coord = tpp_cell.coordinate

                r_vat, _ = _find_label_in_column(ws, "VAT", label_col, search_start, search_end)
                vat_coord = None
                if r_vat and tpp_coord:
                    vat_cell = _first_value_cell_right(ws, r_vat, label_col)
                    vat_cell.value = f"={tpp_coord}/5"
                    vat_cell.number_format = ACC
                    vat_coord = vat_cell.coordinate

                r_opp, _ = _find_label_in_column(ws, "Overall Package Price", label_col, search_start, search_end)
                if not r_opp and r_vat:
                    r_opp = r_vat + 2
                if r_opp and tpp_coord and vat_coord:
                    opp_cell = ws.cell(r_opp, value_col)
                    opp_cell.value = f"={tpp_coord}+{vat_coord}"
                    opp_cell.number_format = ACC

            # ---------- Notes sheet (Q&A, UNWRAPPED) ----------
            notes_ws = _get_notes_ws(wb)
            q1_label = MAIN_NOTES_QUESTIONS[0]
            q2_label = MAIN_NOTES_QUESTIONS[1]
            notes_ws["A2"].value = q1_label
            notes_ws["A3"].value = q2_label
            notes_ws["A2"].alignment = Alignment(wrap_text=False, vertical="top")
            notes_ws["A3"].alignment = Alignment(wrap_text=False, vertical="top")

            def _first_nonempty(colname):
                if colname not in dfp.columns:
                    return ""
                for s in dfp[colname].astype(str).tolist():
                    if s and s.strip() and s.strip().lower() != "nan":
                        return s.strip()
                return ""

            ans1 = _first_nonempty("_note_q1")
            ans2 = _first_nonempty("_note_q2")

            notes_ws["B2"].value = ans1 if ans1 else None
            notes_ws["B3"].value = ans2 if ans2 else None
            notes_ws["B2"].alignment = Alignment(wrap_text=False, vertical="top")
            notes_ws["B3"].alignment = Alignment(wrap_text=False, vertical="top")

            # Save file per provider+contact (prevents overwrites)
            out_bytes = BytesIO()
            wb.save(out_bytes)
            out_bytes.seek(0)
            safe_provider = re.sub(r'[^A-Za-z0-9 _.-]+', '_', prov or "Unknown_Provider")
            safe_contact  = re.sub(r'[^A-Za-z0-9 _.-]+', '_', nm or "Unknown_Contact")
            zf.writestr(f"templates/{safe_provider} - {safe_contact}.xlsx", out_bytes.getvalue())

    zip_buf.seek(0)
    return zip_buf

# ===========================
# Streamlit App (Wishlist only)
# ===========================
st.set_page_config(page_title="MOF Automation", layout="wide")
st.title("MOF Automation")

# Button styles (red submit, green download)
st.markdown("""
<style>
div.stButton > button:first-child { background-color: #d90429 !important; color: white !important; }
div.stDownloadButton > button:first-child { background-color: #2a9d8f !important; color: white !important; }
</style>
""", unsafe_allow_html=True)

st.markdown("""
Upload your files and click **Submit** to get a ZIP containing:

**MOF Templates** → one populated workbook per Provider/Contact  
**Cleaned Data** → your single cleaned dataset
""")

c1, c2, c3 = st.columns(3)
with c1:
    form_file = st.file_uploader("Zoho Forms export (.csv/.xlsx/.xls)", type=["csv","xlsx","xls","txt"], key="form_wishlist")
with c2:
    cost_file = st.file_uploader("MOF Cost Sheet (.csv/.xlsx/.xls)", type=["csv","xlsx","xls","txt"], key="cost_wishlist")
with c3:
    template_file = st.file_uploader("New Template (Excel)", type=["xlsx","xls"], key="tpl_wishlist")

if st.button("Submit", key="submit_wishlist"):
    if not form_file or not cost_file or not template_file:
        st.error("Please upload the Zoho Forms export, MOF Cost Sheet, and the new Template.")
    else:
        with st.spinner("Processing..."):
            try:
                form_df = _read_any_table(form_file, preferred_sheet_name="Form")
                costs_df = _read_any_table(cost_file)
                cleaned_internal = transform_wishlist(form_df, costs_df)  # includes _note_q1/_note_q2
            except Exception as e:
                st.exception(e)
                st.stop()

            # Build results.zip (cleaned export drops hidden notes cols)
            zip_buf = BytesIO()
            with zipfile.ZipFile(zip_buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                cleaned_to_export = cleaned_internal.drop(columns=['_note_q1','_note_q2'], errors='ignore')
                cleaned_to_export = cleaned_to_export.drop(
                    columns=[h for h in NOTES_SOURCE_HEADERS if h in cleaned_to_export.columns],
                    errors='ignore'
                )

                cleaned_bytes = BytesIO()
                cleaned_to_export.to_excel(cleaned_bytes, index=False)
                zf.writestr("data/cleaned_output.xlsx", cleaned_bytes.getvalue())

                try:
                    template_bytes = template_file.read()
                    tpl_zip = _populate_template_bytes(template_bytes, cleaned_internal, costs_df)
                    with zipfile.ZipFile(tpl_zip, 'r') as tplzf:
                        for info in tplzf.infolist():
                            zf.writestr(info.filename, tplzf.read(info.filename))
                except Exception as e:
                    st.exception(RuntimeError(f"Template population failed: {e}"))

            zip_buf.seek(0)
            st.success(f"Done. Cleaned {len(cleaned_to_export)} rows.")

            missing_costs = cleaned_to_export['Cost'].isna().sum()
            if missing_costs > 0:
                st.warning(f"{missing_costs} row(s) have no Cost match by (Type, Product). Ensure both exist in the MOF Cost Sheet.")
                with st.expander("Preview rows missing Cost"):
                    st.dataframe(
                        cleaned_to_export[cleaned_to_export['Cost'].isna()][
                            ['Provider Name','Name','Type','Event Date (if applicable)','Product']
                        ].head(500)
                    )

            st.download_button(
                "Download Now!",
                data=zip_buf.getvalue(),
                file_name="results.zip",
                mime="application/zip",
                key="dl_wishlist"
            )
